import os
import time

from openpyxl import Workbook
from selenium.webdriver import Chrome
from selenium.webdriver import ChromeOptions
from lxml import etree

import properties_read
from svgCrack import getSvgDic, repSvgStr, delemoji
from urllib.parse import quote

chrome_options = ChromeOptions()

ua='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'
chrome_options.add_argument('user-agent=' + ua)

# 修改windows.navigator.webdriver，防机器人识别机制，selenium自动登陆判别机制
chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
# 隐藏"Chrome正在受到自动软件的控制"
chrome_options.add_argument('disable-infobars')

driver = Chrome(chrome_options=chrome_options)
#窗口最大化
driver.maximize_window()
#隐式等待
#driver.set_page_load_timeout (15)

# CDP执行JavaScript 代码  重定义windows.navigator.webdriver的值
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
        Object.defineProperty(navigator, 'webdriver', {
          get: () => undefined
        })
      """
})

#起始页面
start_url = ''
#想要爬取的关键字
key_words = []
#截止日期
end_time = ''
#等待时间
wait_time = 0.0
#读取配置文件
p = properties_read.Properties('config.properties')
try:
    # 爬取开始url
    start_url = p.getProperties('startUrl')
    # 想要爬取的关键字，从配置文件中取得
    temp_words = p.getProperties('crawShops').split(",")
    #转换编码
    for temp_word in temp_words:
        key_words.append(quote(temp_word))
    # 爬取结束时间
    end_time = p.getProperties('endTime')
    # 爬取结束时间
    wait_time = float(p.getProperties('waitTime'))
except:
    print("config文件不正确---退出----")
    SystemExit

driver.get(start_url)
#等待登录
while True:
    try:
        driver.find_element_by_xpath("//div[@class='countDown']")
        break;
    except:
        pass

time.sleep(1)
print("-----------------------------登录成功---------------------------------------")

for key_word in key_words:
    # 跳转到店铺搜索页
    js_url = "window.location.href = 'https://www.dianping.com/search/keyword/1/0_{}'".format(key_word)
    driver.execute_script(js_url)
    time.sleep(1)
    #已经爬取的店
    craw_results = []
    for root, dirs, files in os.walk(os.getcwd() + '\\result'):
        craw_results = files
    # 所有店铺
    all_shop = []
    # 所有店铺url
    all_shop_url = []
    while True:
        links_dom = driver.find_elements_by_xpath("//div[@id='shop-all-list']/ul/li//div[@class='tit']/a[1]")
        page_shops = driver.find_elements_by_xpath("//div[@id='shop-all-list']/ul/li//div[@class='tit']/a[1]/h4")
        i = 0
        for i in range(len(links_dom)):
            # 是否爬取了
            not_craw_flag = True
            for result in craw_results:
                if page_shops[i].text in result:
                    not_craw_flag = False
                    break
            if not_craw_flag:
                # 取得url
                all_shop_url.append(links_dom[i].get_attribute("href"))
                all_shop.append(page_shops[i].text)
        try:
            # 下一页
            driver.find_element_by_xpath("//a[@title='下一页']").click()
        except:
            break
        time.sleep(1)

    print(all_shop)
    print(all_shop_url)

    #all_shop_url = ['http://www.dianping.com/shop/k9pkZa3f0LnOrgxl','http://www.dianping.com/shop/k8mGvY2r4IE3mlMS','http://www.dianping.com/shop/k2YZuJet7Xd5sJ8A']
    for shop_url in all_shop_url:
        # 跳转到店铺评论页面
        js_url = "window.location.href = '{}/review_all?queryType=sortType&&queryVal=latest'".format(shop_url)
        driver.execute_script(js_url)
        # 创建文件对象
        wb = Workbook()
        # 获取第一个sheet
        ws = wb.active
        # 表头
        ws['A1'] = "店铺名"
        ws['B1'] = "评论"
        ws['C1'] = "时间"
        ws['D1'] = "城市"
        # 调整列宽
        ws.column_dimensions['A'].width = 20.0
        ws.column_dimensions['B'].width = 100.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 20.0
        # 获取第一个sheet
        ws = wb.active
        #print(driver.page_source)
        html = etree.HTML(driver.page_source)
        #取得店铺名称
        try:
            shop = html.xpath("//div[@class='review-shop-wrap']//h1[@class='shop-name']/text()")[0]
        except:
            #店铺没有评论
            break

        end_time_flag = False

        #判断页面的加密方式
        def getEncryptionType(etree_html):
            try:
                svgmtsi_class = etree_html.xpath("//div[@class='reviews-items']/ul/li//div[@class='review-words Hide']/svgmtsi[1]/@class")[0]
            except:
                return 'none'
            if svgmtsi_class == "":
                return 'none'
            elif svgmtsi_class == "review":
                return "woff"
            else:
                return "svg"

        while True:
            #到达截止日
            if end_time_flag:
                break
            html = etree.HTML(driver.page_source)
            #判断页面加密方式
            encryptionType = getEncryptionType(html)
            if encryptionType == "svg":
                # 取得svg字典
                svgDic = getSvgDic(driver.page_source)
            elif encryptionType == "woff":
                # 取得woff字典
                svgDic = []
            reviews = html.xpath("//div[@class='reviews-items']/ul/li//div[@class='review-words Hide']|//div[@class='reviews-items']/ul/li//div[@class='review-words']")
            times = html.xpath("//div[@class='reviews-items']/ul/li//div[@class='misc-info clearfix']//span[@class='time']")
            for i in range(len(reviews)):
                if encryptionType == "svg":
                    # 解密评论
                    review = repSvgStr(svgDic, etree.tostring(reviews[i], encoding="utf-8", pretty_print=True,
                                                              method="html").decode('utf-8'))
                elif encryptionType == "woff":
                    # 解密评论
                    review = etree.tostring(reviews[i], encoding="utf-8", pretty_print=True,
                                                              method="html").decode('utf-8')
                else:
                    aa = etree.tostring(reviews[i], encoding="utf-8", pretty_print=True,
                                   method="html").decode('utf-8')
                    bb = etree.HTML(delemoji(aa))
                    cc = bb.xpath("//div[@class='review-words Hide']/text()|//div[@class='review-words']/text()")[0]
                    review = cc.strip()

                #分割日期 筛选出特殊日期形式 例：2015-03-29 更新于2018-10-21 18:32
                time_array = times[i].text.strip().split('更新于')
                if len(time_array) > 0:
                    time_str = time_array[len(time_array)-1]
                else:
                    time_str = ""
                # 日期到了，跳出
                if time_str < end_time:
                    end_time_flag = True
                    break
                print("---------------------------------------")
                print(shop)
                print(review, times[i])
                print(times[i].text.strip())
                # 写入多个单元格
                ws.append([shop,review,times[i].text.strip(),"上海"])
            #下一页
            try:
                driver.find_element_by_xpath("//a[@title='下一页']").click()
            except:
                break
            time.sleep(wait_time)

        # 保存为爬取结果
        dateStr = time.strftime("%Y%m%d", time.localtime())
        wb.save(shop + "的爬取结果_" + dateStr + ".xlsx")




